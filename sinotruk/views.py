import openpyxl
from django.shortcuts import render, HttpResponseRedirect
from django.contrib.auth import authenticate, login as user_login, logout as user_logout
from django.http import FileResponse, JsonResponse
from .models import *
from openpyxl import Workbook
from .forms import *
from .utils import *
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime


def ip_info(addr):
	from urllib.request import urlopen
	from json import load
	url = 'https://ipinfo.io/' + addr + '/json'
	res = urlopen(url)
	data = load(res)

	if 'city' in data.keys():
		return "Place: " + data['city'] + " " + data['region'] + " " + data['loc']
	elif addr == '127.0.0.1':
		return 'SERVER (LOCAL)'
	else:
		return 'not found'


def login_view(request):
	if request.user.is_authenticated:
		return HttpResponseRedirect('/')
	else:

		if request.method == 'POST':

			login = request.POST.get('login')
			password = request.POST.get('password')

			usr = authenticate(request, username=login, password=password)
			if usr is not None:
				user_login(request, usr)
				Activity.objects.create(
					user=usr,
					action='Авторизация',
					ip=request.META['REMOTE_ADDR'],
					place=ip_info(request.META['REMOTE_ADDR'])
				)
				return HttpResponseRedirect('/')

			else:
				return render(request, 'main/login.html', {'error': 'Пользователь с такими параметрами не найден!'})

		else:
			return render(request, 'main/login.html')


def logout_view(request):
	Activity.objects.create(
		user=request.user,
		action='Выход из аккаунта',
		ip=request.META['REMOTE_ADDR'],
		place=ip_info(request.META['REMOTE_ADDR'])
	)

	user_logout(request)
	return HttpResponseRedirect('/login')


def edit_me(request):
	if request.user.role == 'ADMIN':
		if request.method == 'GET':
			return render(request, 'main/edit_admin.html')
		elif request.method == 'POST':

			username = request.POST.get('username')
			password = request.POST.get('password')
			sub_role = request.POST.get('sub_role')
			name = request.POST.get('name')
			surname = request.POST.get('surname')

			error = None

			if request.user.username != username:
				if CustomUser.objects.filter(username=username).exists():
					error = 'Пользователь с таким именем пользователя уже существует!'
				else:

					Activity.objects.create(
						user=request.user,
						action=f'Смена имени пользователя с "{request.user.username}" на "{username}"',
						ip=request.META['REMOTE_ADDR'],
						place=ip_info(request.META['REMOTE_ADDR'])
					)

					request.user.username = username
					request.user.name = name
					request.user.surname = surname
					request.user.sub_role = sub_role
					request.user.custom_set_password(password)

					user = authenticate(request, username=username, password=password)
					if user is not None:
						user_login(request, user)

					return HttpResponseRedirect('/')

			else:
				request.user.name = name
				request.user.surname = surname
				request.user.sub_role = sub_role
				request.user.custom_set_password(password)

				user = authenticate(request, username=username, password=password)
				if user is not None:
					user_login(request, user)

				return render(request, 'main/edit_admin.html', {'error': error})


def home(request):

	if request.user.is_authenticated:

		class Chat:
			def __init__(self, recipient):
				self.recipient = recipient
				self.unread = len([
					x for x in Message.objects.filter(sender=recipient, recipient=request.user, read=False)
				])

		chats = set([
					x.recipient for x in Message.objects.filter(sender=request.user)
				] + [
					x.sender for x in Message.objects.filter(recipient=request.user)
				])

		chats = [Chat(x) for x in chats]

		return render(request, 'main/home.html', {
			'managers_count': CustomUser.objects.filter(role='MANAGER').count(),
			'DEALER_count': CustomUser.objects.filter(role='DEALER').count(),
			'clients_count': CustomUser.objects.filter(role='CLIENT').count(),
			'chats': chats,
			'first_line': request.user.role in ['MANAGER', 'SUPERMANAGER', 'ADMIN'],
			'task_need': request.user.role in ['MANAGER', 'SUPERMANAGER', 'ADMIN'],
			'trash_order': request.user.role in ['DEALER', 'CLIENT']
		})

	return HttpResponseRedirect('/login')


def get_chat(request, user_id):
	user = CustomUser.objects.get(id=user_id)

	messages = [
		x for x in Message.objects.filter(sender=user, recipient=request.user)
	] + [
		x for x in Message.objects.filter(sender=request.user, recipient=user)
	] + [
		x for x in MessageDocument.objects.filter(sender=request.user, recipient=user)
	] + [
		x for x in MessageDocument.objects.filter(sender=user, recipient=request.user)
	]

	for x in Message.objects.filter(sender=user, recipient=request.user):
		x.read = True
		x.save()

	#for x in MessageDocument.objects.filter(sender=user, recipient=request.user):
	#	x.read = True
	#	x.save()

	class MsgView:
		def __init__(self, msg):
			if isinstance(msg, Message):
				self.type = 'text'
			else:
				self.type = 'blob'
			self.obj = msg

	return render(request, 'main/chat.html', {
		'messages': [MsgView(x) for x in sorted(messages, key=lambda x: x.date)],
		'recipient': user,
	})


@csrf_exempt
def get_user_details(request, user_id):
	if request.user.is_authenticated:
		user = CustomUser.objects.get(id=user_id)

		return JsonResponse({
			'username': user.username,
		}, status=200)


@csrf_exempt
def delete_user(request, user_id):
	if request.user.is_authenticated:
		user = CustomUser.objects.get(id=user_id)
		user_name = user.username
		user.delete()

		Activity.objects.create(
			user=request.user,
			action=f'Удаление пользователя {user_name}',
			ip=request.META['REMOTE_ADDR'],
			place=ip_info(request.META['REMOTE_ADDR'])
		)

		return JsonResponse({}, status=200)


def managers(request):
	managers = CustomUser.objects.filter(role='MANAGER')
	supermanagers = CustomUser.objects.filter(role='SUPERMANAGER')

	return render(request, 'main/managers.html', {
		'managers': managers,
		'manager_count': managers.count() + supermanagers.count(),
		'supermanagers': supermanagers
	})


def add_manager(request):
	if request.user.role == 'ADMIN':
		if request.method == 'POST':

			username = request.POST.get('username')
			password = request.POST.get('password')
			name = request.POST.get('name')
			surname = request.POST.get('surname')

			user = CustomUser.objects.create_user(username=username, password=password, name=name, surname=surname, role='MANAGER')
			user.custom_set_password(password)
			user.save()

			Activity.objects.create(
				user=request.user,
				action=f'Добавление менеджера {user.username}',
				ip=request.META['REMOTE_ADDR'],
				place=ip_info(request.META['REMOTE_ADDR'])
			)

			return HttpResponseRedirect('/managers')

		else:
			return render(request, 'main/add_manager.html')
	else:
		return HttpResponseRedirect('/')


def dealer(request):
	managers = CustomUser.objects.filter(role='DEALER')

	return render(request, 'main/dealers.html', {
		'DEALER': managers,
		'DEALER_count': managers.count(),
	})


def add_dealer(request):
	if request.user.role == 'ADMIN':
		if request.method == 'POST':

			username = request.POST.get('username')
			password = request.POST.get('password')
			firm = request.POST.get('firm')

			try:
				user = CustomUser.objects.create(username=username, password=password, role='DEALER', dealer_name=firm)
				user.custom_set_password(password)
				user.role = 'DEALER'
				user.save()

				Activity.objects.create(
					user=request.user,
					action=f'Добавление дилера {user.username}',
					ip=request.META['REMOTE_ADDR'],
					place=ip_info(request.META['REMOTE_ADDR'])
				)
			except:
				return render(request, 'main/add_dealer.html', {'error': 'Ошибка создания пользователя, пользователь с таким именем пользователя уже существует!'})

			return HttpResponseRedirect('/dealers')

		else:
			return render(request, 'main/add_dealer.html')
	else:
		return HttpResponseRedirect('/')


def clients(request):
	clients = CustomUser.objects.filter(role='CLIENT')

	return render(request, 'main/clients.html', {
		'clients': clients,
		'clients_count': clients.count()
	})


def add_client(request):
	if request.user.role == 'ADMIN':
		if request.method == 'POST':

			username = request.POST.get('username')
			name = request.POST.get('name')
			surname = request.POST.get('surname')
			password = request.POST.get('clear_password')

			user = CustomUser.objects.create_user(username=username, name=name, surname=surname, role='CLIENT', password=password)
			user.custom_set_password(password)
			user.save()

			Activity.objects.create(
				user=request.user,
				action=f'Добавление клиента {user.username}',
				ip=request.META['REMOTE_ADDR'],
				place=ip_info(request.META['REMOTE_ADDR'])
			)

			return HttpResponseRedirect('/clients')

		else:
			return render(request, 'main/add_client.html')
	else:
		return HttpResponseRedirect('/')


def pricelist(request):
	act = Activity.objects.filter(action='Обновление прайс-листа').order_by('-time')
	update_date = None

	if len(act):
		update_date = act.last().time

	class pw:
		def __init__(self, p):
			self.product = p
			self.saved = ProductOnTrash.objects.filter(user=request.user, product=p).count() > 0

	if request.method == 'GET':
		products = [pw(x) for x in Product.objects.all()]

		staff = request.user.role in ['ADMIN', 'MANAGER', 'SUPERMANAGER']

		return render(request, 'main/pricelist.html', {
			'update_date': update_date,
			'products': products,
			'staff': staff
		})

	elif request.method == 'POST':
		nom = request.POST.get('nom')
		nom = nom.lower() if nom else None
		char = request.POST.get('char')
		char = char.lower() if char else None
		mark = request.POST.get('mark')
		mark = mark.lower() if mark else None

		products = [x for x in Product.objects.all()]

		if nom:
			products = [
				x for x in products if nom in str(x.serial_number).lower()
			]
		if char:
			products = [
				x for x in products if char in str(x.name).lower()
			]
		if mark:
			products = [
				x for x in products if mark in str(x.mark).lower()
			]

		products = [pw(x) for x in products]
		staff = request.user.role in ['ADMIN', 'MANAGER', 'SUPERMANAGER']

		return render(request, 'main/pricelist.html', {
			'update_date': update_date,
			'products': products,
			'n': nom, 'c': char, 'm': mark,
			'staff': staff
		})


def add_pricelist(request):
	if request.method == "POST":

		f = request.FILES["excel_file"]
		end = request.FILES["excel_file"].name.split(".")[-1]

		if end in ['xlsx', 'xls']:

			with open(f'{settings.BASE_DIR}/media/pricelist/file.{end}', 'wb+') as destination:
				for chunk in f.chunks():
					destination.write(chunk)

			fix_excel(f'{settings.BASE_DIR}/media/pricelist/file.{end}')

			excel = openpyxl.load_workbook(f'{settings.BASE_DIR}/media/pricelist/file.{end}')
			sheet = excel.active

			rows = []
			for r in sheet.rows:
				rows.append([])
				for cell in r:
					rows[-1].append(cell.value)

			start_index = 0
			for i in range(len(rows)):
				if 'Номенклатура.Артикул' in [str(x).strip() for x in rows[i] if x]:
					start_index = i
					break

			dt = [str(x).strip() for x in rows[start_index]]
			sn = dt.index('Номенклатура.Артикул')
			name = dt.index('Ценовая группа/ Номенклатура')
			ost = dt.index('Остаток')
			mrk = dt.index('Марки')
			prc = dt.index('Дилер')

			Product.objects.all().delete()
			for row in rows[start_index+1:]:
				sn_ = row[sn]
				if sn_:
					name_ = row[name]
					ost_ = str(row[ost])
					if ost_.isdigit():
						ost_ = int(ost_)

					mrk_ = row[mrk]
					prc_ = str(row[prc]).replace('руб.', '').strip()
					if prc_.replace('.', '').isdigit():
						prc_ = float(prc_)

					Product.objects.create(
						serial_number=sn_,
						name=name_,
						count=ost_,
						price=prc_,
						mark=mrk_
					)

			Activity.objects.create(
				user=request.user,
				action=f'Обновление прайс-листа',
				ip=request.META['REMOTE_ADDR'],
				place=ip_info(request.META['REMOTE_ADDR'])
			)

			return HttpResponseRedirect('/pricelist')
		else:
			return render(request, 'main/add_pricelist.html', {
				'error': 'Недопустимый тип файла!'
			})
	return render(request, 'main/add_pricelist.html')


def add_folder(request):
	if request.method == "POST":
		form = AddFolderForm(request.POST)

		if form.is_valid():

			Folder.objects.create(name=request.POST.get('folder-name'))

			Activity.objects.create(
				user=request.user,
				action=f'Создание папки {request.POST.get("folder-name")}',
				ip=request.META['REMOTE_ADDR'],
				place=ip_info(request.META['REMOTE_ADDR'])
			)

			return HttpResponseRedirect('/files')

		else:
			return render(request, 'main/add_folder.html', {'form': form})

	return render(request, 'main/add_folder.html', {'form': AddFolderForm()})


def folder_detail(request, fid):
	folder = Folder.objects.get(id=fid)
	files_ = Document.objects.filter(folder=folder)

	Activity.objects.create(
		user=request.user,
		action=f'Просмотр папки {folder.name}',
		ip=request.META['REMOTE_ADDR'],
		place=ip_info(request.META['REMOTE_ADDR'])
	)

	back_link = '/files'
	if folder.parent_folder:
		back_link = f'/folders/{folder.parent_folder.id}'

	return render(request, 'main/folder_detail.html', {
		'files': files_,
		'folder': folder,
		'folders': Folder.objects.filter(parent_folder=folder),
		'back_link': back_link,
		'staff': request.user.role not in ['DEALER', 'CLIENT'],
		'can_delete': request.user.role not in ['DEALER', 'CLIENT']
	})


def files(request):
	folders = Folder.objects.filter(parent_folder=None)
	files_ = Document.objects.filter(folder=None)
	staff = request.user.role in ['ADMIN', 'MANAGER', 'SUPERMANAGER']

	return render(request, 'main/files.html', {
		'files': files_,
		'folders': folders,
		'files_count': files_.count(),
		'folders_count': folders.count(),
		'staff': staff
	})


def add_file(request):
	if request.method == 'POST':

		form = AddFileForm(request.POST, request.FILES)
		if form.is_valid():
			file = form.save()
			if file.document.path.endswith('.pdf'):
				file.file_type = 'pdf'
			if file.document.path.endswith('.doc'):
				file.file_type = 'doc'
			if file.document.path.endswith('.docx'):
				file.file_type = 'doc'
			if file.document.path.endswith('.xlsx'):
				file.file_type = 'xlsx'
			if file.document.path.split('.')[-1] in ['jpg', 'jpeg', 'png', 'gif']:
				file.file_type = 'image'

			file.save()

			Activity.objects.create(
				user=request.user,
				ip=request.META['REMOTE_ADDR'],
				action=f'Загрузка файла "{file.title}"',
				place=ip_info(request.META['REMOTE_ADDR'])
			)

			return HttpResponseRedirect('/files')
	else:
		return render(request, 'main/add_file.html', {'form': AddFileForm()})


@csrf_exempt
def update_user_task(request):
	if request.user.is_authenticated:
		if request.method == 'POST':

			request.user.sub_role = request.POST.get('task')
			request.user.save()

			return JsonResponse({}, status=200)

		else:
			return JsonResponse({}, status=500)
	else:
		return JsonResponse({}, status=500)


@csrf_exempt
def update_user_receiver(request):
	if request.user.is_authenticated:
		if request.user.role == 'MANAGER':
			if request.method == 'POST':

				request.user.receive_emails = {
					'true': True,
					'false': False
				}[request.POST.get('value')]
				request.user.save()

				return JsonResponse({}, status=200)

			else:
				return JsonResponse({}, status=500)
		else:
			return JsonResponse({}, status=500)
	else:
		return JsonResponse({}, status=500)


def add_folder_to_folder(request, folder_id):
	if request.user.is_authenticated:

		if request.method == 'POST':

			fname = request.POST.get('folder-name')

			f = Folder.objects.create(name=fname, parent_folder=Folder.objects.get(id=folder_id))
			Activity.objects.create(
				user=request.user,
				ip=request.META['REMOTE_ADDR'],
				action=f'Создание папки "{fname}"',
				place=ip_info(request.META['REMOTE_ADDR'])
			)

			return HttpResponseRedirect('/folders/' + str(f.id))

		if request.method == 'GET':
			return render(request, 'main/add_folder.html', {
				'folder': Folder.objects.get(id=folder_id)
			})

	else:
		return HttpResponseRedirect('/login')


def add_file_to_folder(request, folder_id):
	if request.user.is_authenticated:

		if request.method == 'POST':

			form = AddFileForm(request.POST, request.FILES)
			if form.is_valid():
				file = form.save()
				if file.document.path.endswith('.pdf'):
					file.file_type = 'pdf'
				if file.document.path.endswith('.doc'):
					file.file_type = 'doc'
				if file.document.path.endswith('.docx'):
					file.file_type = 'doc'
				if file.document.path.endswith('.xlsx'):
					file.file_type = 'xlsx'
				if file.document.path.split('.')[-1] in ['jpg', 'jpeg', 'png', 'gif']:
					file.file_type = 'image'

				file.folder = Folder.objects.get(id=folder_id)
				file.save()

				Activity.objects.create(
					user=request.user,
					ip=request.META['REMOTE_ADDR'],
					action=f'Загрузка файла "{file.title}"',
					place=ip_info(request.META['REMOTE_ADDR'])
				)

				return HttpResponseRedirect('/folders/' + str(folder_id))
		else:
			return render(request, 'main/add_file.html', {'form': AddFileForm()})

	else:
		return HttpResponseRedirect('/login')


def get_document(request, doc_id):
	file = Document.objects.get(id=doc_id)

	Activity.objects.create(
		user=request.user,
		ip=request.META['REMOTE_ADDR'],
		action=f'Скачивание файла "{file.title}"',
		place=ip_info(request.META['REMOTE_ADDR'])
	)

	return FileResponse(file.document)


def activity(request):
	managers = CustomUser.objects.filter(role='MANAGER')
	DEALER = CustomUser.objects.filter(role='DEALER')
	clients = CustomUser.objects.filter(role='CLIENT')
	supermanagers = CustomUser.objects.filter(role='SUPERMANAGER')
	admins = CustomUser.objects.filter(role='ADMIN')

	return render(request, 'main/activity.html', {
		'managers': managers,
		'DEALER': DEALER,
		'clients': clients,
		'supermanagers': supermanagers,
		'admins': admins,
		'managers_count': managers.count(),
		'DEALER_count': DEALER.count(),
		'clients_count': clients.count(),
		'supermanagers_count': supermanagers.count(),
		'admins_count': admins.count()
	})


def load_activity(request):
	if request.method == "GET":
		managers = CustomUser.objects.filter(role='MANAGER')
		DEALER = CustomUser.objects.filter(role='DEALER')
		clients = CustomUser.objects.filter(role='CLIENT')
		supermanagers = CustomUser.objects.filter(role='SUPERMANAGER')
		admins = CustomUser.objects.filter(role='ADMIN')

		return render(request, 'main/load_activity.html', {
			'managers': managers,
			'DEALER': DEALER,
			'clients': clients,
			'supermanagers': supermanagers,
			'admins': admins,
			'managers_count': managers.count(),
			'DEALER_count': DEALER.count(),
			'clients_count': clients.count(),
			'supermanagers_count': supermanagers.count(),
			'admins_count': admins.count()
		})


def edit_user(request, user_id):
	user = CustomUser.objects.get(id=user_id)

	if request.method == 'GET':
		if user.role == 'MANAGER':
			if request.user.role == 'ADMIN':
				form = EditManagerForm(instance=user)
				return render(request, 'main/edit_user.html', {'user': user, 'form': form})
			if request.user.role == 'SUPERMANAGER':
				form = EditSuperManagerForm(instance=user)
				return render(request, 'main/edit_user.html', {'user': user, 'form': form})
		elif user.role == 'DEALER':
			form = EditDealerForm(instance=user)
			return render(request, 'main/edit_user.html', {'user': user, 'form': form})
		elif user.role == 'CLIENT':
			form = EditClientForm(instance=user)
			return render(request, 'main/edit_user.html', {'user': user, 'form': form})
		elif user.role == 'SUPERMANAGER':
			form = EditManagerForm(instance=user)
			return render(request, 'main/edit_user.html', {'user': user, 'form': form})
		else:
			return HttpResponseRedirect('/')

	elif request.method == 'POST':

		Activity.objects.create(
			user=request.user,
			action=f'Изменение данных пользователя {user.username}',
			ip=request.META['REMOTE_ADDR'],
			place=ip_info(request.META['REMOTE_ADDR'])
		)

		if user.role == 'MANAGER':
			form = EditManagerForm(instance=user, data=request.POST)

			if form.is_valid():

				user.username = form.cleaned_data['username']
				user.name = form.cleaned_data['name']
				user.surname = form.cleaned_data['surname']
				user.role = form.cleaned_data['role']
				user.sub_role = form.cleaned_data['sub_role']
				user.custom_set_password(form.cleaned_data['clear_password'])

				return HttpResponseRedirect('/managers')

			else:
				return render(request, 'main/edit_user.html', {'user': user, 'form': form})

		elif user.role == 'DEALER':
			form = EditDealerForm(instance=user, data=request.POST)

			if form.is_valid():

				user.username = form.cleaned_data['username']
				user.name = form.cleaned_data['name']
				user.surname = form.cleaned_data['surname']
				user.role = form.cleaned_data['role']
				user.dealer_name = form.cleaned_data['dealer_name']
				user.custom_set_password(form.cleaned_data['clear_password'])

				return HttpResponseRedirect('/DEALER')

			else:
				return render(request, 'main/edit_user.html', {'user': user, 'form': form})

		else:
			return HttpResponseRedirect('/')


def user_history(request, user_id):
	user = CustomUser.objects.get(id=user_id)
	history = Activity.objects.filter(user=user).order_by('-id')

	return render(request, 'main/user_history.html', {
		'user': user,
		'history': history,
		'hcount': history.count()
	})


@csrf_exempt
def add_message(request):
	if request.user.is_authenticated:
		if request.method == 'POST':

			message = Message.objects.create(
				sender=CustomUser.objects.get(id=int(request.POST.get('from'))),
				recipient=CustomUser.objects.get(id=int(request.POST.get('to'))),
				text=request.POST.get('text')
			)

			return JsonResponse({
				'from': message.sender.username,
				'to': message.recipient.username,
				'text': message.text,
				'date': message.date.strftime('%d %B %Y г. %H:%M')
			}, status=200)


def send_file_message(request, uid):
	recipient = CustomUser.objects.get(id=uid)

	if request.method == 'POST':
		form = FileUploadForm(request.POST, request.FILES)

		if form.is_valid():

			msg = form.save(commit=False)
			msg.recipient = recipient
			msg.sender = request.user
			msg.save()

			return HttpResponseRedirect(f'/chat/{uid}')
		else:

			return render(request, 'main/send_file_message.html', {
				'form': form,
				'recipient': recipient
			})

	return render(request, 'main/send_file_message.html', {
		'form': FileUploadForm(),
		'recipient': recipient
	})


@csrf_exempt
def load_activity_file(request):
	if request.method == 'POST':

		data = {}

		if request.POST['start'] and request.POST['end']:
			start = datetime.strptime(request.POST['start'], '%Y-%m-%d').date()
			end = datetime.strptime(request.POST['end'], '%Y-%m-%d').date()

			for el in request.POST:
				if request.POST[el].isdigit():
					user = CustomUser.objects.get(id=int(request.POST[el]))

					if user in data:
						data[user] += [x for x in Activity.objects.filter(user=user, time__gte=start, time__lte=end)]
					else:
						data[user] = [x for x in Activity.objects.filter(user=user, time__gte=start, time__lte=end)]

		else:
			for el in request.POST:
				if request.POST[el].isdigit():
					user = CustomUser.objects.get(id=int(request.POST[el]))

					if user in data:
						data[user] += [x for x in Activity.objects.filter(user=user)]
					else:
						data[user] = [x for x in Activity.objects.filter(user=user)]

		wb = Workbook()
		ws1 = wb.active
		ws1.title = "Дилеры"
		ws2 = wb.create_sheet("Клиенты")
		ws3 = wb.create_sheet("Менеджеры")
		ws4 = wb.create_sheet("Суперменеджеры")
		ws5 = wb.create_sheet("Админы")

		for user in data:
			if len(data[user]):
				sheet = {
					'DEALER': ws1,
					'CLIENT': ws2,
					'MANAGER': ws3,
					'SUPERMANAGER': ws4,
					'ADMIN': ws5
				}[user.role]
				sheet.append([user.username])
				sheet.append([])

				for act in data[user]:
					sheet.append(['', str(act.time), str(act.ip), str(act.place), str(act.action)])

				sheet.append([])
				sheet.append([])
				sheet.append([])

		filename = settings.BASE_DIR / 'media/documents/activity.xlsx'
		os.remove(filename)
		wb.save(filename=filename)

		return JsonResponse({'link': '/media/documents/activity.xlsx'}, status=200)


def trash(request):
	if request.user.is_authenticated:

		products = ProductOnTrash.objects.filter(user=request.user)

		return render(request, 'main/trash.html', {
			'products': products,
			'pcount': products.count()
		})

	else:
		return HttpResponseRedirect('/login')


def add_product_to_trash(request, pid):
	if request.user.is_authenticated:

		if ProductOnTrash.objects.filter(user=request.user, product_id=pid).exists():
			pass
		else:
			prod = Product.objects.get(id=pid)
			ProductOnTrash.objects.create(user=request.user, product=prod)
		return JsonResponse({}, status=200)

	else:
		return HttpResponseRedirect('/login')


@csrf_exempt
def remove_from_trash(request):
	if request.method == 'POST':

		ps = ProductOnTrash.objects.filter(
			user=request.user,
			product__id=int(request.POST.get('pid'))
		)

		if ps.exists():
			ps.delete()
			return JsonResponse({}, status=200)
		else:
			return JsonResponse({}, status=404)

	else:
		return JsonResponse({}, status=500)


@csrf_exempt
def remove_from_trash_by_pid(request, pid):

	ps = ProductOnTrash.objects.filter(
		user=request.user,
		product__id=pid
	)

	if ps.exists():
		ps.delete()
		return JsonResponse({}, status=200)
	else:
		return JsonResponse({}, status=404)


@csrf_exempt
def add_order(request):
	if request.method == 'POST':

		prods = []

		for el in request.POST:
			prods.append(
				(Product.objects.get(id=int(el.replace('p', ''))), int(request.POST[el]))
			)

		order = Order.objects.create(user=request.user)

		for el in prods:
			OrderItem.objects.create(order=order, product=el[0], count=el[1])

		send_emails_to_managers(order)

		ProductOnTrash.objects.filter(user=request.user).delete()

		return JsonResponse({}, status=200)

	else:
		return JsonResponse({}, status=500)


def my_orders(request):
	if request.user.is_authenticated:

		#class OrderView:
		#	def __init__(self, order):
		#		self.order = order
		#		self.items = OrderItem.objects.filter(order=order)
		#
		#orders = [OrderView(x) for x in Order.objects.filter(user=request.user)]

		return render(request, 'main/my_orders.html', {
			'orders': Order.objects.filter(user=request.user).order_by('-id')
		})

	return HttpResponseRedirect('/login')


def my_order_detail(request, order_id):
	if request.user.is_authenticated:

		order = Order.objects.get(id=order_id)

		if order.user == request.user:
			items = OrderItem.objects.filter(order=order)

			total_price = sum([x.count * x.product.price for x in items])

			return render(request, 'main/order_detail.html', {
				'order': order,
				'items': items,
				'tp': total_price
			})
		else:
			return HttpResponseRedirect('/')

	return HttpResponseRedirect('/login')


def order_detail(request, order_id):
	if request.user.is_authenticated:

		order = Order.objects.get(id=order_id)
		items = OrderItem.objects.filter(order=order)
		total_price = sum([x.count * x.product.price for x in items])

		return render(request, 'main/order_detail.html', {
			'order': order,
			'items': items,
			'tp': total_price
		})

	return HttpResponseRedirect('/login')


def main_orders(request):
	if request.user.is_authenticated:

		class OrderView:
			def __init__(self, order):
				self.order = order
				self.total_price = sum([
					(x.product.price * x.count) for x in OrderItem.objects.filter(order=order)
				])

		orders = [OrderView(x) for x in Order.objects.all()]

		return render(request, 'main/orders.html', {
			'orders': orders
		})

	else:
		HttpResponseRedirect('/login')


def contacts(request):
	if request.user.is_authenticated:

		users = CustomUser.objects.all().order_by('role')

		return render(request, 'main/contacts.html', {
			"users": users
		})

	else:
		return HttpResponseRedirect('/login')


def activity_stats(request):

	users = CustomUser.objects.all().order_by('role').exclude(role__in=['ADMIN'])

	class UserLogin:
		def __init__(self, user):
			self.user = user
			self.count = Activity.objects.filter(user=user, action='Авторизация').count()

	class UserMessageCounter:
		def __init__(self, user):
			self.user = user
			self.count = Message.objects.filter(sender=user).count() + MessageDocument.objects.filter(sender=user).count()

	class UserOrder:
		def __init__(self, user):
			self.user = user
			self.count = Order.objects.filter(user=user).count()

	data = []
	msgs = []
	orders = []
	for user in users:
		data.append(UserLogin(user))
		msgs.append(UserMessageCounter(user))
		orders.append(UserOrder(user))

	return render(request, 'main/activity_stats.html', {
		'logins': data,
		'messages': msgs,
		'orders': orders
	})


def delete_folder(request, fid):
	if request.user.is_authenticated:
		if request.method == 'GET':

			folder = Folder.objects.get(id=fid)
			parent = folder.parent_folder
			folder.delete()

			if parent:
				return HttpResponseRedirect(f'/folders/{parent.id}')
			return HttpResponseRedirect('files')


def make_order_as_complete(request, order_id):
	if request.user.is_authenticated:
		if request.user.role in ['MANAGER', 'SUPERMANAGER', 'ADMIN']:

			order = Order.objects.get(id=order_id)
			order.complete = True
			order.save()

			return HttpResponseRedirect(f'/orders')

	return HttpResponseRedirect('/')


def delete_document(request, doc_id):
	if request.user.is_authenticated:

		doc = Document.objects.get(id=doc_id)
		link = None

		if doc.folder:
			link = f'/folders/{doc.folder.id}'
		else:
			link = '/files'

		doc.delete()

		return HttpResponseRedirect(link)
